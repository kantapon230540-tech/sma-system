"""Strip another site's data out of the two solidification export templates.

`rebuild_export_templates.py` cannot do this job: its solid sources (Att-2 / Att-3) live
in ~/Downloads and are no longer there, and its `cell_plan_solid` only clears cells the
row map already knows about. Three kinds of residue sit outside that map and therefore
shipped in every export:

  1. `Data` — a peer-comparison matrix holding other Bridgestone plants' 1-5 solidification
     scores by name. 432 literal scores in the DP template; the Safety template's numeric
     block was already cleared, so only its site names remain.
  2. Role-sheet Judge columns of the TWIN half. Safety has no twin map at all, so 136 of
     BMT's DP judgements were never cleared and are never overwritten; DP keeps 9 at rows
     its twin map does not reach.
  3. Cover identity — site code, assessment dates, issue date.

Everything is spliced as BYTES via _patch_rows; re-serialising sheet XML with ElementTree
makes the workbook unopenable (see ref_xlsx_namespace_trap).

Idempotent: running it twice is a no-op. Run from the repo root:
    python3 tools/sanitise_solid_templates.py [--dry-run]
"""
import re, sys, zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import openpyxl
from export.checksheet import _patch_rows, _sheet_paths, _split   # noqa: E402

OUT = Path(__file__).resolve().parent.parent / "export_templates"

JUDGEMENTS = {"yes", "no", "n/a", "na", "not rolled out yet"}
# Plant codes appearing as column headers on the Data sheets and on the covers.
SITE_CODE = re.compile(r"^(BM[TU]|BS[A-Za-z]+|TMOT|ACA|TBSC[A-Za-z]*|佐賀|日モ)$")
# Free-text identity on the cover sheets: site, the assessment dates, the issue date.
COVER_CELLS = {"D4", "J4", "X1", "W2"}


def judge_columns(ws):
    """Columns headed 'Judge' on this sheet. The role sheets do not share a layout —
    the header sits on row 3 or row 6 and the column is H or J — so it is read, not assumed."""
    cols = set()
    for row in ws.iter_rows(min_row=1, max_row=12):
        for c in row:
            if isinstance(c.value, str) and c.value.strip().lower() == "judge":
                cols.add(c.column_letter)
    return cols


def plan_for(path):
    """{sheet: {row: {col: None}}} plus a human-readable tally of what is being removed."""
    wb = openpyxl.load_workbook(path)
    plan, tally = {}, {}

    def clear(sheet, ref):
        col, row = _split(ref)
        plan.setdefault(sheet, {}).setdefault(row, {})[col] = None

    for sn in wb.sheetnames:
        ws = wb[sn]

        # 2. leftover judgements in any Judge column, on EITHER half
        n = 0
        for col in judge_columns(ws):
            for r in range(1, ws.max_row + 1):
                v = ws[f"{col}{r}"].value
                if isinstance(v, str) and v.strip().lower() in JUDGEMENTS:
                    clear(sn, f"{col}{r}"); n += 1
        if n: tally[f"{sn}: judgements"] = n

        # 3. cover identity
        if sn.strip().lower().endswith("cover"):
            n = 0
            for ref in COVER_CELLS:
                if ws[ref].value not in (None, ""):
                    clear(sn, ref); n += 1
            if n: tally[f"{sn}: identity cells"] = n

        # 1. the peer-comparison matrix
        if sn == "Data":
            names = scores = 0
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if v is None or (isinstance(v, str) and v.startswith("=")):
                        continue
                    if isinstance(v, str) and SITE_CODE.match(v.strip()):
                        clear(sn, c.coordinate); names += 1
                    elif isinstance(v, (int, float)) and 1 <= v <= 5 and c.column >= 5:
                        clear(sn, c.coordinate); scores += 1
            if names: tally["Data: site names"] = names
            if scores: tally["Data: peer scores"] = scores

    wb.close()
    return plan, tally


def apply(path, plan):
    with zipfile.ZipFile(path) as z:
        paths = _sheet_paths(z)
        patched = {}
        for sheet, edits in plan.items():
            xml, _, miss = _patch_rows(z.read(paths[sheet]), edits)
            assert not miss, f"{sheet}: rows not found {miss}"
            patched[paths[sheet]] = xml
        items = [(i, patched.get(i.filename) or z.read(i.filename)) for i in z.infolist()]
    tmp = path.with_suffix(".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as w:
        for item, data in items:
            w.writestr(item, data)
    tmp.replace(path)


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    for name in ("safety_solid_checksheet.xlsx", "dp_solid_checksheet.xlsx"):
        path = OUT / name
        plan, tally = plan_for(path)
        total = sum(len(v) for rows in plan.values() for v in rows.values())
        print(f"\n{name}  ({path.stat().st_size / 1024:.0f} KB)")
        for k, v in sorted(tally.items()):
            print(f"   {k}: {v}")
        print(f"   -> {total} cells")
        if not dry and total:
            apply(path, plan)
            print(f"   written ({path.stat().st_size / 1024:.0f} KB)")
