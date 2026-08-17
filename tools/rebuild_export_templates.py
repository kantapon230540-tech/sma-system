"""Build the export templates from the ORIGINAL sources, sanitised in one reproducible pass.

Att-1 is TBSCN's completed 2025 result. Their data hides in EIGHT places, found one at a
time over several rounds — this script does all of them so the process is repeatable:

  1. Judgement / Reason cells                     (the obvious one)
  2. Remark and "F-up result" columns             — never cleared before; AB78 still read
                                                     "Same issue from QM audit"
  3. Cover Sheet identity + injury/TIR/SIR data
  4. Embedded pictures                            — 8 wrap-up slides, "Confidential S2",
                                                     3.77 MB, anchored beside Reason
  5. Saved AutoFilter criteria                    — <filter val="..."> keeps the Reason text
                                                     of every row that was filtered on
  6. sharedStrings orphans                        — clearing a cell drops the reference,
                                                     not the string
  7. Chart caches                                 — radars redraw the previous site's scores
  8. docProps + comment authors, and the x15ac:absPath SharePoint URL, which names the
     TBSCN audit folder

Everything is spliced as BYTES — re-serialising with ElementTree breaks the namespaces and
Excel rejects the file (see ref_xlsx_namespace_trap).
"""
import json, re, shutil, sys, zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import openpyxl
from export.checksheet import _patch_rows, _sheet_paths           # noqa: E402
from _sanitise import (blank_orphan_strings, zero_chart_cache,      # noqa: E402
                      referenced_shared_strings)

DL = Path("/Users/ktp/Downloads")
OUT = Path(__file__).resolve().parent.parent / "export_templates"
# The MFG master is the controlled blank sheet SE09-PR-01-1-G rev5, vendored under
# export_templates/sources/ so a rebuild does not depend on ~/Downloads surviving.
# It replaced Att-1, which was TBSCN's COMPLETED result and leaked their data eight ways.
SRC = {
    "mfg":    Path(__file__).resolve().parent.parent / "export_templates" / "sources" /
              "SE09-PR-01-1-G_Appendix B Safety Maturity Assessment Sheet_rev5.xlsx",
    "safety": DL / "Att-2 Safety Solidification Assessment Sheet (Rev.0 Dec-2025).xlsx",
    "dp":     DL / "Att-3 Disaster Prevention audit check sheet.xlsx",
    # The warehouse master, vendored the same way. Content-identical to the sheet the
    # 159-question bank was built from (verified question-by-question: same No mapping,
    # same judgement criteria on all 159), but it still carried the author's local
    # absPath and docProps names, so it goes through the same sanitising pass.
    "wh":     Path(__file__).resolve().parent.parent / "export_templates" / "sources" /
              "Safety maturity assessment checklist for WH_Oct 2024.xlsx",
}

# No column, then Judgement / Reason / Remark / F-up. System Checklist is shifted +1.
MFG_SHEETS = {
    "Leadership Checklist":          ("F", ["Y", "Z", "AA", "AB"]),
    "Teammate Engagement Checklist": ("F", ["Y", "Z", "AA", "AB"]),
    "Organization Checklist":        ("F", ["Y", "Z", "AA", "AB"]),
    "System Checklist":              ("G", ["Z", "AA", "AB", "AC"]),
}
MFG_COVER = {
    6: {"D": None, "H": None, "J": None},
    9: {"B": None}, 10: {"B": None}, 11: {"B": None},
    15: {"J": None},
    17: {"B": None, "D": None, "F": None, "H": None, "J": None},
    21: {"F": None, "H": None, "J": None},
    22: {"F": None, "H": None, "J": None},
    23: {"F": None, "H": None, "J": None},
}
DASH = {"(Ref.)MA Dashboard": {4: {"F": None}},
        "(Ref.)MA Dashboard (DP_divide)": {4: {"F": None}}}

# Warehouse. No Cover Sheet — the single (Ref.)Dashboard carries the identity block
# (Site F4 / Date F5 / Assessor F6 / Assessment type F7), and there is no Remark or
# F-up column, so Judgement + Reason are all a completed sheet would hold.
WH_SHEETS = {
    "Leadership":    ("D", ["K", "L"]),
    "TM Engagement": ("D", ["K", "L"]),
    "Organization":  ("D", ["K", "L"]),
    "System":        ("E", ["L", "M"]),
}
WH_DASH = {"(Ref.)Dashboard": {4: {"F": None}, 5: {"F": None}, 6: {"F": None}}}


def cell_plan_mfg():
    wb = openpyxl.load_workbook(SRC["mfg"], read_only=True, data_only=True)
    plan = {"Cover Sheet": dict(MFG_COVER), **{k: dict(v) for k, v in DASH.items()}}
    for sheet, (nocol, cols) in MFG_SHEETS.items():
        ws, edits = wb[sheet], {}
        ncol = openpyxl.utils.column_index_from_string(nocol)
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                             max_col=32, values_only=True), 1):
            v = row[ncol - 1]
            if v is None:
                continue
            if not str(v).strip().replace(".0", "").isdigit():
                continue
            edits[i] = {c: None for c in cols}
        plan[sheet] = edits
    wb.close()
    return plan


def cell_plan_wh():
    """Same shape as cell_plan_mfg, for the warehouse master. F7 (assessment type) is
    deliberately NOT cleared — it is the master's own default and the exporter overwrites
    it from the record's `kind`, so a sheet exported without one still reads sensibly."""
    wb = openpyxl.load_workbook(SRC["wh"], read_only=True, data_only=True)
    plan = {k: dict(v) for k, v in WH_DASH.items()}
    for sheet, (nocol, cols) in WH_SHEETS.items():
        ws, edits = wb[sheet], {}
        ncol = openpyxl.utils.column_index_from_string(nocol)
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                             max_col=20, values_only=True), 1):
            v = row[ncol - 1]
            if v is None:
                continue
            if not str(v).strip().replace(".0", "").isdigit():
                continue
            edits[i] = {c: None for c in cols}
        plan[sheet] = edits
    wb.close()
    return plan


def cell_plan_solid(track):
    maps = json.loads((OUT / "ssdpma_solid_maps.json").read_text())[track]
    plan = {}
    def add(m):
        # Only clear cells that are genuinely INPUTS. On the DP-shaped halves the overall
        # sheet's Judge cells are formulas reading the role sheets (=OP!H4) — blanking those
        # destroys the workbook's wiring, which is exactly what an earlier pass did.
        if m["overall_is_input"]:
            for r in m["rows"].values():
                plan.setdefault(m["overall_sheet"], {}).setdefault(int(r), {}).update(
                    {c: None for c in m["judge_cols"]})
        for sh, col, r in m["role_targets"].values():
            plan.setdefault(sh, {}).setdefault(int(r), {})[col] = None
        for ref in m.get("direct", {}).values():
            plan.setdefault(m["overall_sheet"], {}).setdefault(
                int(ref[1:]), {})[ref[0]] = None
    add(maps)
    if maps.get("twin"):
        add(maps["twin"])
    return plan


def strip_hyperlinks(sheet_xml, rels_xml):
    """Remove external hyperlinks. The System Checklist linked each evidence document on
    SharePoint, so the targets spell out the previous site's audit folder
    (.../AUDIT/SMA/TBSCn Recover/1. TBSCn Verify Genba result/...). The <hyperlink> elements
    and their relationships both have to go, or the workbook has dangling rIds."""
    n = len(re.findall(rb"<hyperlink\b", sheet_xml))
    sheet_xml = re.sub(rb"<hyperlinks>.*?</hyperlinks>", b"", sheet_xml, flags=re.S)
    sheet_xml = re.sub(rb"<hyperlink\b[^>]*/>", b"", sheet_xml)
    if rels_xml is not None:
        # drop EVERY hyperlink relationship, not just the ones a <hyperlink> still cited —
        # Excel leaves orphaned hyperlink rels behind, and those carry the URL just the same
        rels_xml, k = re.subn(
            rb'<Relationship\b(?=[^>]*Type="[^"]*/hyperlink")[^>]*/>', b"", rels_xml)
        n = max(n, k)
    return sheet_xml, rels_xml, n


def strip_filters(xml):
    """Drop saved AutoFilter criteria — <filter val="..."> holds the Reason text of every
    row that was filtered on — and turn filterMode off. The <autoFilter ref> itself stays
    so the _FilterDatabase defined name keeps resolving."""
    n = len(re.findall(rb"<filter\s", xml))
    xml = re.sub(rb"<autoFilter\b([^>]*?)>.*?</autoFilter>", rb"<autoFilter\1/>", xml, flags=re.S)
    xml = re.sub(rb'(<sheetPr\b[^>]*?)\sfilterMode="1"', rb"\1", xml)
    return xml, n


def build(kind, template_name, plan, drop_drawing_on=None):
    src = SRC[kind]
    z = zipfile.ZipFile(src)
    paths = _sheet_paths(z)
    patched, notes = {}, {}

    # 1-3. cells
    cleared = 0
    for sheet, edits in plan.items():
        p = paths[sheet]
        xml, _, miss = _patch_rows(z.read(p), edits)
        assert not miss, f"{sheet}: rows not found {miss}"
        patched[p] = xml
        cleared += sum(len(v) for v in edits.values())
    notes["cells cleared"] = cleared

    # 5. autofilter criteria on every sheet
    filt = 0
    for n in z.namelist():
        if not re.match(r"xl/worksheets/sheet\d+\.xml$", n):
            continue
        base = patched.get(n) or z.read(n)
        base, c = strip_filters(base)
        if c or base != (patched.get(n) or z.read(n)):
            patched[n] = base
        filt += c
    notes["autofilter criteria dropped"] = filt

    # 5b. external hyperlinks (evidence links naming the previous site's folders)
    links = 0
    for n in list(z.namelist()):
        if not re.match(r"xl/worksheets/sheet\d+\.xml$", n):
            continue
        rp = n.replace("worksheets/", "worksheets/_rels/") + ".rels"
        rels = (patched.get(rp) or (z.read(rp) if rp in z.namelist() else None))
        sx, rx, c = strip_hyperlinks(patched.get(n) or z.read(n), rels)
        if c:
            patched[n] = sx
            if rx is not None:
                patched[rp] = rx
            links += c
    notes["external hyperlinks removed"] = links

    # 4a. unreferenced junk parts. The warehouse master carries two "[trash]/NNNN.dat"
    # blocks — FF FF FF FF then padding, stored uncompressed with a 1980 timestamp, the
    # signature of a zip repair pass. No content type declares them and no rels reference
    # them, so they are not OOXML parts at all; they would otherwise ride into every export.
    drop = {n for n in z.namelist() if n.startswith("[trash]/")}
    if drop:
        notes["junk parts dropped"] = sorted(drop)

    # 4. evidence pictures
    if drop_drawing_on:
        sp = paths[drop_drawing_on]
        rp = sp.replace("worksheets/", "worksheets/_rels/") + ".rels"
        # start from the ALREADY-patched rels — the hyperlink strip ran first, and re-reading
        # the original here silently threw its work away
        rels = (patched.get(rp) or z.read(rp)).decode("utf-8")
        m = re.search(r'<Relationship[^>]*Id="([^"]+)"[^>]*Type="[^"]*/drawing"[^>]*'
                      r'Target="([^"]+)"[^>]*/>', rels) or \
            re.search(r'<Relationship[^>]*Type="[^"]*/drawing"[^>]*Id="([^"]+)"[^>]*'
                      r'Target="([^"]+)"[^>]*/>', rels)
        if m:
            rid, dpath = m.group(1), "xl/" + m.group(2).replace("../", "")
            drels = dpath.replace("drawings/", "drawings/_rels/") + ".rels"
            imgs = set()
            if drels in z.namelist():
                imgs = {"xl/" + t.replace("../", "") for t in
                        re.findall(r'Type="[^"]*/image"[^>]*Target="([^"]+)"',
                                   z.read(drels).decode("utf-8"))}
            keep = set()
            for other in z.namelist():
                if re.match(r"xl/drawings/_rels/drawing\d+\.xml\.rels$", other) and other != drels:
                    keep |= {"xl/" + t.replace("../", "") for t in
                             re.findall(r'Type="[^"]*/image"[^>]*Target="([^"]+)"',
                                        z.read(other).decode("utf-8"))}
            imgs -= keep
            drop |= {dpath, drels} | imgs
            patched[sp] = re.sub(rb"<drawing[^>]*/>", b"", patched.get(sp) or z.read(sp), count=1)
            patched[rp] = re.sub(r'<Relationship[^>]*Id="%s"[^>]*/>' % re.escape(rid), "",
                                 rels, count=1).encode("utf-8")
            ct = z.read("[Content_Types].xml").decode("utf-8")
            patched["[Content_Types].xml"] = re.sub(
                r'<Override[^>]*PartName="/%s"[^>]*/>' % re.escape(dpath), "", ct, count=1
            ).encode("utf-8")
            notes["pictures removed"] = f"{len(imgs)} ({sum(z.getinfo(i).file_size for i in imgs):,} bytes)"

    # 8. absPath (SharePoint URL naming the source audit folder) + fullCalcOnLoad
    wbx = (z.read("xl/workbook.xml")).decode("utf-8")
    had_path = "absPath" in wbx
    wbx = re.sub(r"<x15ac:absPath[^>]*/>", "", wbx)
    if "fullCalcOnLoad" not in wbx:
        wbx = re.sub(r"<calcPr([^>]*?)/>", r'<calcPr\1 fullCalcOnLoad="1"/>', wbx, count=1)
    patched["xl/workbook.xml"] = wbx.encode("utf-8")
    notes["absPath removed"] = had_path

    out = OUT / template_name
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as w:
        for item in z.infolist():
            if item.filename in drop:
                continue
            w.writestr(item, patched.get(item.filename) or z.read(item.filename))
    z.close()

    # 6-8. orphan strings, chart caches, docProps, comment authors
    zz = zipfile.ZipFile(out)
    used = referenced_shared_strings(zz)
    p2, ch = {}, 0
    if "xl/sharedStrings.xml" in zz.namelist():
        xml, blanked, total = blank_orphan_strings(zz.read("xl/sharedStrings.xml"), used)
        p2["xl/sharedStrings.xml"] = xml
        notes["sharedStrings blanked"] = f"{blanked}/{total}"
    for n in zz.namelist():
        if re.match(r"xl/charts/chart\d+\.xml$", n):
            xml, c = zero_chart_cache(zz.read(n))
            if c: p2[n] = xml; ch += c
    notes["chart points zeroed"] = ch
    if "docProps/core.xml" in zz.namelist():
        x = zz.read("docProps/core.xml")
        names = re.findall(rb"<dc:creator>([^<]*)</dc:creator>", x) + \
                re.findall(rb"<cp:lastModifiedBy>([^<]*)</cp:lastModifiedBy>", x)
        x = re.sub(rb"<dc:creator>[^<]*</dc:creator>", b"<dc:creator></dc:creator>", x)
        x = re.sub(rb"<cp:lastModifiedBy>[^<]*</cp:lastModifiedBy>",
                   b"<cp:lastModifiedBy></cp:lastModifiedBy>", x)
        p2["docProps/core.xml"] = x
        notes["names removed"] = [n.decode() for n in names if n]
    for n in zz.namelist():
        if re.match(r"xl/comments\d+\.xml$", n):
            p2[n] = re.sub(rb"<author>[^<]*</author>", b"<author></author>", zz.read(n))
    tmp = out.with_suffix(".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as w:
        for item in zz.infolist():
            w.writestr(item, p2.get(item.filename) or zz.read(item.filename))
    zz.close(); tmp.replace(out)

    print(f"\n{template_name}  ({out.stat().st_size/1024:.0f} KB)")
    for k, v in notes.items():
        print(f"   {k}: {v}")


if __name__ == "__main__":
    build("mfg", "mfg_checksheet.xlsx", cell_plan_mfg(),
          drop_drawing_on="System Checklist")
    build("safety", "safety_solid_checksheet.xlsx", cell_plan_solid("safety_solid"))
    build("dp", "dp_solid_checksheet.xlsx", cell_plan_solid("dp_solid"))
    build("wh", "warehouse_checksheet.xlsx", cell_plan_wh())
